#!/usr/bin/env python3
"""
Utility to fix clinical trial date labels in protocol Excel files.

Scans all subdirectories in source_data/protocols for .xlsx files and updates
the "study" sheet column E entries:
  - "Sponsor Approval Date" -> "Approval Date"
  - "Protocol Effective Date" -> "Effective Date"
"""

import os
from pathlib import Path
from openpyxl import load_workbook


def fix_excel_file(filepath: Path) -> bool:
    """
    Fix date labels in the study sheet of an Excel file.

    Returns True if any changes were made, False otherwise.
    """
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        print(f"  Error loading {filepath}: {e}")
        return False

    if "study" not in wb.sheetnames:
        print(f"  No 'study' sheet in {filepath.name}")
        return False

    ws = wb["study"]
    changes_made = False

    for row in ws.iter_rows(min_col=5, max_col=5):
        cell = row[0]
        if cell.value == "Sponsor Approval Date":
            cell.value = "Approval Date"
            changes_made = True
            print(f"  Changed 'Sponsor Approval Date' -> 'Approval Date' at row {cell.row}")
        elif cell.value == "Protocol Effective Date":
            cell.value = "Effective Date"
            changes_made = True
            print(f"  Changed 'Protocol Effective Date' -> 'Effective Date' at row {cell.row}")

    if changes_made:
        wb.save(filepath)
        print(f"  Saved {filepath.name}")
    else:
        print(f"  No changes needed in {filepath.name}")

    wb.close()
    return changes_made


def main():
    script_dir = Path(__file__).parent
    protocols_dir = script_dir / "source_data" / "protocols"

    if not protocols_dir.exists():
        print(f"Error: {protocols_dir} does not exist")
        return

    total_files = 0
    files_modified = 0

    for subdir in protocols_dir.iterdir():
        if not subdir.is_dir():
            continue

        xlsx_files = list(subdir.glob("*.xlsx"))
        if not xlsx_files:
            continue

        for xlsx_file in xlsx_files:
            # Skip temporary Excel files
            if xlsx_file.name.startswith("~$"):
                continue

            print(f"Processing: {subdir.name}/{xlsx_file.name}")
            total_files += 1

            if fix_excel_file(xlsx_file):
                files_modified += 1

    print(f"\nSummary: Processed {total_files} files, modified {files_modified}")


if __name__ == "__main__":
    main()
